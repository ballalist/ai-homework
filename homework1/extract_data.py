import openpyxl
import json
import re
from datetime import datetime

base = r'\\wsl.localhost\Ubuntu\home\bond\projects\ai-homework-test'

wb_data = openpyxl.load_workbook(f'{base}\\player_feedback_300_dataset (1).xlsx')
ws = wb_data['Feedback_Raw']

# Read all rows
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
print('Header:', header)
print(f'Total data rows: {len(rows)-1}')

# Print all rows as JSON
data = []
for row in rows[1:]:
    r = dict(zip(header, row))
    if r.get('date') and hasattr(r['date'], 'strftime'):
        r['date'] = r['date'].strftime('%Y-%m-%d')
    data.append(r)

with open(f'{base}\\feedback_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('Saved feedback_data.json')
print(f'Sample (first 5):')
for d in data[:5]:
    print(d)

# Read category guide
ws_cat = wb_data['Category_Guide']
cats = list(ws_cat.iter_rows(values_only=True))
print('\nCategory Guide:')
for c in cats:
    print(c)
