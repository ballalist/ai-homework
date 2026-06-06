import openpyxl
import json

base = r'\\wsl.localhost\Ubuntu\home\bond\projects\ai-homework-test'

# Read Report file
wb_report = openpyxl.load_workbook(f'{base}\\AI Workflow Challenge Player Feedback Insight Report.xlsx')
print('=== REPORT FILE ===')
print('Sheets:', wb_report.sheetnames)
for sheet in wb_report.sheetnames:
    ws = wb_report[sheet]
    print(f'\n--- Sheet: {sheet} ({ws.max_row} rows x {ws.max_column} cols) ---')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(cell is not None for cell in row):
            print(f'  Row {i+1}: {row}')

print('\n\n=== DATASET FILE ===')
wb_data = openpyxl.load_workbook(f'{base}\\player_feedback_300_dataset (1).xlsx')
print('Sheets:', wb_data.sheetnames)
for sheet in wb_data.sheetnames:
    ws = wb_data[sheet]
    print(f'\n--- Sheet: {sheet} ({ws.max_row} rows x {ws.max_column} cols) ---')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True)):
        print(f'  Row {i+1}: {row}')
    print(f'  ... (total {ws.max_row} rows)')
