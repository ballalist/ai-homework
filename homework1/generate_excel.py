import json
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

base = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(base, 'classified_feedback.json'), encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()

# ===== Sheet 1: Feedback Analysis =====
ws = wb.active
ws.title = 'Feedback_Analysis'

HEADERS = [
    'Feedback ID', 'Date', 'Source', 'Player ID', 'Player Segment',
    'Platform', 'Game Version', 'Game Area', 'Player Feedback',
    'Sentiment', 'Category', 'Priority', 'AI Summary', 'Suggested Owner'
]
FIELDS = [
    'feedback_id', 'date', 'source', 'player_id', 'player_segment',
    'platform', 'game_version', 'game_area_hint', 'player_feedback',
    'sentiment', 'category', 'priority', 'ai_summary', 'suggested_owner'
]

# --- Header row ---
header_fill = PatternFill('solid', fgColor='4F46E5')
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border

ws.row_dimensions[1].height = 28

# --- Colour maps ---
SENT_FILL = {
    'Positive': PatternFill('solid', fgColor='D1FAE5'),
    'Neutral':  PatternFill('solid', fgColor='F1F5F9'),
    'Negative': PatternFill('solid', fgColor='FEE2E2'),
}
SENT_FONT = {
    'Positive': Font(name='Calibri', color='065F46', size=10),
    'Neutral':  Font(name='Calibri', color='475569', size=10),
    'Negative': Font(name='Calibri', color='991B1B', size=10),
}
PRI_FILL = {
    'High':   PatternFill('solid', fgColor='FEE2E2'),
    'Medium': PatternFill('solid', fgColor='FEF3C7'),
    'Low':    PatternFill('solid', fgColor='D1FAE5'),
}
PRI_FONT = {
    'High':   Font(name='Calibri', bold=True, color='991B1B', size=10),
    'Medium': Font(name='Calibri', bold=True, color='92400E', size=10),
    'Low':    Font(name='Calibri', bold=True, color='065F46', size=10),
}
default_font = Font(name='Calibri', size=10)
default_align = Alignment(vertical='top', wrap_text=False)
wrap_align   = Alignment(vertical='top', wrap_text=True)
alt_fill = PatternFill('solid', fgColor='F8FAFF')

# --- Data rows ---
for row_idx, d in enumerate(data, 2):
    alt = (row_idx % 2 == 0)
    for col_idx, field in enumerate(FIELDS, 1):
        val = d.get(field, '')
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border

        sent = d.get('sentiment', '')
        pri  = d.get('priority', '')

        if field == 'sentiment':
            cell.fill = SENT_FILL.get(sent, PatternFill())
            cell.font = SENT_FONT.get(sent, default_font)
            cell.alignment = Alignment(horizontal='center', vertical='top')
        elif field == 'priority':
            cell.fill = PRI_FILL.get(pri, PatternFill())
            cell.font = PRI_FONT.get(pri, default_font)
            cell.alignment = Alignment(horizontal='center', vertical='top')
        elif field in ('player_feedback', 'ai_summary'):
            cell.font = default_font
            cell.alignment = wrap_align
        else:
            cell.font = default_font
            cell.alignment = default_align
            if alt:
                cell.fill = alt_fill

    ws.row_dimensions[row_idx].height = 36

# --- Column widths ---
COL_WIDTHS = [12, 12, 18, 12, 16, 14, 14, 14, 55, 12, 22, 10, 55, 24]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = ws.dimensions

# ===== Sheet 2: Summary Stats =====
ws2 = wb.create_sheet('Summary_Stats')

from collections import Counter
sentiments = Counter(d['sentiment'] for d in data)
categories = Counter(d['category'] for d in data)
priorities = Counter(d['priority'] for d in data)
sources    = Counter(d['source'] for d in data)
segments   = Counter(d['player_segment'] for d in data)
platforms  = Counter(d['platform'] for d in data)
total = len(data)

h2_fill = PatternFill('solid', fgColor='7C3AED')
h2_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
section_fill = PatternFill('solid', fgColor='EDE9FE')
section_font = Font(name='Calibri', bold=True, color='4C1D95', size=10)
val_font = Font(name='Calibri', size=10)
val_align = Alignment(horizontal='left', vertical='center')

def write_section(ws2, start_row, title, counter, total):
    # Section title
    hcell = ws2.cell(row=start_row, column=1, value=title)
    hcell.fill = h2_fill
    hcell.font = h2_font
    hcell.alignment = Alignment(horizontal='left', vertical='center')
    ws2.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    ws2.row_dimensions[start_row].height = 22
    start_row += 1
    # Sub-header
    for col, txt in enumerate(['Label', 'Count', '% of Total'], 1):
        c = ws2.cell(row=start_row, column=col, value=txt)
        c.fill = section_fill
        c.font = section_font
        c.alignment = Alignment(horizontal='center')
    ws2.row_dimensions[start_row].height = 18
    start_row += 1
    for label, cnt in counter.most_common():
        ws2.cell(row=start_row, column=1, value=label).font = val_font
        ws2.cell(row=start_row, column=2, value=cnt).font = val_font
        pct_cell = ws2.cell(row=start_row, column=3, value=round(cnt/total*100, 1))
        pct_cell.font = val_font
        pct_cell.number_format = '0.0"%"'
        ws2.row_dimensions[start_row].height = 16
        start_row += 1
    return start_row + 1

row = 1
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 14

row = write_section(ws2, row, 'Sentiment',  sentiments, total)
row = write_section(ws2, row, 'Category',   categories, total)
row = write_section(ws2, row, 'Priority',   priorities, total)
row = write_section(ws2, row, 'Source',     sources,    total)
row = write_section(ws2, row, 'Segment',    segments,   total)
row = write_section(ws2, row, 'Platform',   platforms,  total)

# ===== Save =====
out_path = os.path.join(base, 'feedback_analysis.xlsx')
wb.save(out_path)
print(f'Excel saved: {out_path}')
print(f'  Feedback_Analysis: {total} rows')
print(f'  Summary_Stats: sentiment / category / priority / source / segment / platform')
